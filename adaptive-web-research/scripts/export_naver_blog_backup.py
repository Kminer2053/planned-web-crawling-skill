#!/usr/bin/env python3

import csv
import json
import re
import sys
from collections import Counter
from datetime import datetime
from html import unescape
from html.parser import HTMLParser
from pathlib import Path
from urllib.parse import urlparse


NAVER_HOST_SUFFIXES = (
    "naver.com",
    "blog.naver.com",
    "m.blog.naver.com",
    "section.blog.naver.com",
    "nid.naver.com",
    "help.naver.com",
    "admin.blog.naver.com",
    "blogimgs.pstatic.net",
    "ssl.pstatic.net",
    "blogthumb.pstatic.net",
    "postfiles.pstatic.net",
)


class LinkExtractor(HTMLParser):
    def __init__(self):
        super().__init__()
        self.links = []
        self._current_href = None
        self._current_text = []

    def handle_starttag(self, tag, attrs):
        if tag.lower() == "a":
            attrs_dict = dict(attrs)
            href = attrs_dict.get("href", "")
            if href.startswith(("http://", "https://")):
                self._current_href = href
                self._current_text = []

    def handle_data(self, data):
        if self._current_href is not None:
            self._current_text.append(data)

    def handle_endtag(self, tag):
        if tag.lower() == "a" and self._current_href is not None:
            self.links.append(
                {
                    "url": self._current_href.strip(),
                    "anchor_text": clean_space(" ".join(self._current_text)),
                }
            )
            self._current_href = None
            self._current_text = []


def clean_space(text: str) -> str:
    text = unescape(text or "")
    text = text.replace("\xa0", " ")
    text = re.sub(r"[ \t\r\f\v]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def strip_tags(html: str) -> str:
    html = re.sub(r"(?is)<script.*?>.*?</script>", " ", html)
    html = re.sub(r"(?is)<style.*?>.*?</style>", " ", html)
    html = re.sub(r"(?is)<noscript.*?>.*?</noscript>", " ", html)
    text = re.sub(r"(?i)<br\s*/?>", "\n", html)
    text = re.sub(r"(?is)<[^>]+>", "\n", text)
    text = unescape(text)
    text = re.sub(r"[ \t\r\f\v]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def regex_first(pattern: str, text: str) -> str:
    match = re.search(pattern, text, flags=re.I | re.S)
    return clean_space(match.group(1)) if match else ""


def parse_datetime(raw: str):
    raw = clean_space(raw)
    if not raw:
        return "", ""
    for fmt in ("%Y. %m. %d. %H:%M", "%Y. %m. %d."):
        try:
            dt = datetime.strptime(raw, fmt)
            return dt.date().isoformat(), dt.isoformat(timespec="minutes")
        except ValueError:
            continue
    return raw, raw


def extract_content_html(html: str) -> str:
    for pattern in [
        r'(<div class="se-main-container".*?</div>\s*</div>\s*</div>)',
        r'(<div id="postViewArea".*?</div>\s*</div>)',
        r'(<div class="post-view".*?</div>\s*</div>)',
    ]:
        match = re.search(pattern, html, flags=re.I | re.S)
        if match:
            return match.group(1)
    return html


def extract_hashtags(content_html: str) -> list[str]:
    tags = []
    for tag in re.findall(r'<span class="__se-hash-tag">(.*?)</span>', content_html, flags=re.I | re.S):
        cleaned = clean_space(tag)
        if cleaned:
            tags.append(cleaned)
    seen = set()
    deduped = []
    for tag in tags:
        if tag not in seen:
            deduped.append(tag)
            seen.add(tag)
    return deduped


def normalize_url(url: str) -> str:
    return re.sub(r"\?utm_[^#]+", "", (url or "").strip())


def is_external_url(url: str) -> bool:
    if not url.startswith(("http://", "https://")):
        return False
    host = (urlparse(url).hostname or "").lower()
    if not host:
        return False
    return not any(host == suffix or host.endswith("." + suffix) for suffix in NAVER_HOST_SUFFIXES)


def extract_external_links(content_html: str, content_text: str, raw_html: str) -> list[dict]:
    parser = LinkExtractor()
    parser.feed(content_html)

    links = []
    seen = set()
    for item in parser.links:
        href = normalize_url(item["url"])
        if not is_external_url(href):
            continue
        if href in seen:
            continue
        seen.add(href)
        host = (urlparse(href).hostname or "").lower()
        links.append({"url": href, "domain": host, "anchor_text": item["anchor_text"]})

    for href in re.findall(r"https?://[^\s<>\"]+", "\n".join([content_text, raw_html])):
        href = normalize_url(href.rstrip(").,]"))
        if not is_external_url(href):
            continue
        if href in seen:
            continue
        seen.add(href)
        host = (urlparse(href).hostname or "").lower()
        links.append({"url": href, "domain": host, "anchor_text": href})
    return links


def extract_content_text(content_html: str) -> str:
    text = strip_tags(content_html)
    for marker in ["\n공감\n", "\n댓글\n", "\n이 블로그\n", "\n활동정보\n", "\n글 보내기 서비스 안내\n"]:
        pos = text.find(marker)
        if pos != -1:
            text = text[:pos].strip()
    return text


def infer_post_type(title: str, series_name: str, content_text: str) -> str:
    base = " ".join([title or "", series_name or "", content_text[:300]])
    if "뉴스클리핑" in base:
        return "news-clipping"
    if "오늘의 로코노미" in base:
        return "daily-brief"
    if "특집리뷰" in base:
        return "special-review"
    if "인사이트" in base:
        return "insight"
    if "주말축제" in base:
        return "festival-roundup"
    return "article"


def summarize_links(links: list[dict]) -> tuple[str, str]:
    domains = [link["domain"] for link in links if link["domain"]]
    top_domains = [f"{domain}({count})" for domain, count in Counter(domains).most_common()]
    return "; ".join(top_domains), "\n".join(link["url"] for link in links)


def export_backup(backup_dir: Path):
    manifest = json.loads((backup_dir / "manifest.json").read_text(encoding="utf-8"))
    exports_dir = backup_dir / "exports"
    exports_dir.mkdir(exist_ok=True)

    post_rows = []
    link_rows = []

    for post in manifest["posts"]:
        raw_path = Path(post["rawPath"])
        html = raw_path.read_text(encoding="utf-8", errors="ignore")
        content_html = extract_content_html(html)

        title = regex_first(r'<meta property="og:title" content="([^"]+)"', html) or post.get("resolvedTitle") or post.get("title") or ""
        author = regex_first(r'<span class="nick">.*?<a [^>]*>(.*?)</a>', html) or regex_first(r'<span class="nick">(.*?)</span>', html)
        series_name = regex_first(r'<div class="blog2_series">.*?<a [^>]*class="pcol2"[^>]*>(.*?)</a>', html)
        published_raw = regex_first(r'<span class="se_publishDate">(.*?)</span>', html)
        published_date, published_at = parse_datetime(published_raw)
        content_text = extract_content_text(content_html)
        hashtags = extract_hashtags(content_html)
        external_links = extract_external_links(content_html, content_text, html)
        external_domain_summary, external_url_list = summarize_links(external_links)

        lines = [line.strip() for line in content_text.splitlines() if line.strip()]
        title_clean = title.replace("[공지]", "").replace(" : 네이버 블로그", "").strip()
        filtered_lines = [line for line in lines if line not in {title, title_clean, author, published_raw, "[공지]"}]
        summary = filtered_lines[0] if filtered_lines else ""
        is_notice = "[공지]" in title or title.startswith("공지")
        post_type = infer_post_type(title_clean, series_name, content_text)

        post_rows.append(
            {
                "log_no": post["logNo"],
                "post_url": post.get("url", ""),
                "title": title_clean,
                "is_notice": int(bool(is_notice)),
                "post_type": post_type,
                "series_name": series_name,
                "author": author,
                "published_date": published_date,
                "published_at": published_at,
                "summary": summary,
                "content_text": content_text,
                "content_char_count": len(content_text),
                "content_line_count": len(filtered_lines),
                "hashtag_count": len(hashtags),
                "hashtags": ", ".join(hashtags),
                "external_link_count": len(external_links),
                "external_domains": external_domain_summary,
                "external_urls": external_url_list,
                "raw_html_path": str(raw_path),
                "text_path": post.get("textPath", ""),
            }
        )

        for idx, link in enumerate(external_links, start=1):
            link_rows.append(
                {
                    "log_no": post["logNo"],
                    "post_title": title_clean,
                    "link_index": idx,
                    "domain": link["domain"],
                    "url": link["url"],
                    "anchor_text": link["anchor_text"],
                }
            )

    post_rows.sort(key=lambda row: row["published_at"] or row["published_date"], reverse=True)
    link_rows.sort(key=lambda row: (row["log_no"], row["link_index"]))

    posts_csv = exports_dir / "posts.csv"
    links_csv = exports_dir / "links.csv"

    with posts_csv.open("w", encoding="utf-8-sig", newline="") as fp:
        writer = csv.DictWriter(fp, fieldnames=list(post_rows[0].keys()))
        writer.writeheader()
        writer.writerows(post_rows)

    with links_csv.open("w", encoding="utf-8-sig", newline="") as fp:
        writer = csv.DictWriter(fp, fieldnames=list(link_rows[0].keys()))
        writer.writeheader()
        writer.writerows(link_rows)

    workbook_path = exports_dir / "blog_backup.xlsx"
    workbook_written = False
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws_posts = wb.active
        ws_posts.title = "posts"
        ws_posts.append(list(post_rows[0].keys()))
        for row in post_rows:
            ws_posts.append(list(row.values()))

        ws_links = wb.create_sheet("links")
        ws_links.append(list(link_rows[0].keys()))
        for row in link_rows:
            ws_links.append(list(row.values()))

        for ws in [ws_posts, ws_links]:
            for idx, column in enumerate(ws.columns, start=1):
                max_len = max(len(str(cell.value or "")) for cell in column[:100])
                ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 60)
            ws.freeze_panes = "A2"

        wb.save(workbook_path)
        workbook_written = True
    except Exception:
        workbook_written = False

    return {
        "backup_dir": str(backup_dir),
        "posts_csv": str(posts_csv),
        "links_csv": str(links_csv),
        "xlsx": str(workbook_path) if workbook_written else "",
        "post_count": len(post_rows),
        "link_count": len(link_rows),
    }


def main():
    if len(sys.argv) != 2:
        print("Usage: export_naver_blog_backup.py <backup_dir>", file=sys.stderr)
        sys.exit(1)
    result = export_backup(Path(sys.argv[1]))
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
